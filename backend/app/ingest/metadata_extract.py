"""Извлечение структурированных метаданных (ISA-Tab / Allotrope-подобные) из текста."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from app.config import settings
from app.models import DocumentMetadata, IsaTabRecord


def _relative_source(path: Path) -> str:
    try:
        return str(path.relative_to(settings.data_dir_path))
    except ValueError:
        return str(path)


def _extract_key_values(text: str, patterns: list[tuple[str, str]]) -> dict[str, Any]:
    found: dict[str, Any] = {}
    for key, pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            found[key] = match.group(1).strip()
    return found


def _extract_docx_meta(path: Path, meta: DocumentMetadata) -> DocumentMetadata:
    try:
        from docx import Document

        doc = Document(str(path))
        core = doc.core_properties
        if core.author and core.author not in meta.authors:
            meta.authors.append(core.author)
        if core.last_modified_by and core.last_modified_by not in meta.authors:
            meta.authors.append(core.last_modified_by)
        if core.created:
            meta.date = meta.date or core.created.strftime("%Y-%m-%d")
        if core.modified and not meta.date:
            meta.date = core.modified.strftime("%Y-%m-%d")
    except Exception:
        pass
    return meta


def _parse_xlsx_isa_tab(path: Path) -> IsaTabRecord | None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None

    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
    except Exception:
        return None

    sheet = wb.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))[:30]
    if len(rows) < 2:
        return None

    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    factors: dict[str, Any] = {}
    measurements: dict[str, Any] = {}
    for row in rows[1:]:
        for i, cell in enumerate(row):
            if cell is None or i >= len(header) or not header[i]:
                continue
            h = header[i]
            if any(k in h for k in ("ph", "доз", "темп", "реагент")):
                factors[header[i]] = cell
            if any(k in h for k in ("извлеч", "recover", "выход", "содерж")):
                measurements[header[i]] = cell

    if not factors and not measurements:
        return None

    first_row = {header[i]: rows[1][i] for i in range(min(len(header), len(rows[1]))) if header[i]}
    return IsaTabRecord(
        investigation_id=path.stem,
        study_id=f"{path.stem}_study",
        assay_id=f"{path.stem}_assay",
        factor_names=list(factors.keys())[:8],
        factor_values=factors,
        measurement_type=next(iter(measurements.keys()), "assay_result"),
        measurement_value=str(next(iter(measurements.values()), "")),
        unit="%" if measurements else None,
        raw_sample_characteristics=first_row,
    )


def _build_allotrope_uri(material: str | None, process: str = "flotation") -> str:
    mat = (material or "ore").replace(" ", "-")
    return f"allotrope:process/{process}/{mat}-laboratory-assay"


def enrich_metadata_from_content(
    path: Path,
    content: str,
    metadata: DocumentMetadata | None = None,
) -> DocumentMetadata:
    """Дополняет метаданные параметрами эксперимента из XLSX/DOCX/PDF."""
    meta = metadata or DocumentMetadata()
    if not meta.title:
        meta.title = path.name
    if not meta.source:
        meta.source = _relative_source(path)

    suffix = path.suffix.lower()
    if suffix == ".docx":
        meta = _extract_docx_meta(path, meta)

    date_match = re.search(
        r"(?:дата|date)\s*[:=]?\s*(\d{4}[-./]\d{1,2}[-./]\d{1,2}|\d{1,2}\.\d{1,2}\.\d{2,4})",
        content,
        re.I,
    )
    if date_match and not meta.date:
        meta.date = date_match.group(1)

    author_match = re.findall(
        r"(?:автор|authors?|исследователь)\s*[:=]\s*([^\n;]{3,80})",
        content,
        re.I,
    )
    for a in author_match[:3]:
        if a.strip() not in meta.authors:
            meta.authors.append(a.strip())

    if suffix in {".xlsx", ".xls"}:
        meta.tags.append("structured_experiment")
        meta.material_type = meta.material_type or _guess_material_type(content)
        meta.sample_id = meta.sample_id or _extract_sample_id(content)
        isa = _parse_xlsx_isa_tab(path)
        if isa:
            meta.isa_tab = isa

    params = _extract_key_values(
        content,
        [
            ("pH", r"pH\s*[:=]?\s*(\d+(?:[.,]\d+)?(?:\s*[-–—]\s*\d+(?:[.,]\d+)?)?)"),
            ("temperature_C", r"температур[аы]?\s*[:=]?\s*(\d+(?:[.,]\d+)?)"),
            ("reagent_dosage", r"(\d+(?:[.,]\d+)?)\s*(?:кг|г)\s*/?\s*т"),
            (
                "recovery_pct",
                r"(?:извлекаем\w+\s+металл\s+в\s+хвостах|извлечени[ея])\s*[:=]?\s*(\d+(?:[.,]\d+)?)\s*%",
            ),
            ("grade_pct", r"содержани[ея]\s*[:=]?\s*(\d+(?:[.,]\d+)?)\s*%"),
        ],
    )
    if params:
        meta.process_parameters.update(params)
        meta.experiment_conditions.update(params)

    measurements = _extract_key_values(
        content,
        [
            ("Cu_recovery", r"Cu[^\d]*(\d+(?:[.,]\d+)?)\s*%"),
            ("yield", r"выход\s*[:=]?\s*(\d+(?:[.,]\d+)?)\s*%"),
        ],
    )
    if measurements:
        meta.measurement_results.update(measurements)

    if meta.isa_tab and meta.isa_tab.factor_values:
        meta.process_parameters.update(
            {k: v for k, v in meta.isa_tab.factor_values.items() if v is not None}
        )
        if meta.isa_tab.measurement_value:
            meta.measurement_results.setdefault(
                meta.isa_tab.measurement_type or "assay_result",
                meta.isa_tab.measurement_value,
            )

    if "лаборатор" in content.lower():
        meta.instrument = meta.instrument or "лабораторная флотационная установка"

    steps = re.findall(
        r"(?:шаг|этап|step)\s*\d+[:\.]?\s*([^\n]{10,120})",
        content,
        re.IGNORECASE,
    )
    if steps:
        meta.protocol_steps = steps[:8]

    meta.allotrope_process_uri = meta.allotrope_process_uri or _build_allotrope_uri(
        meta.material_type
    )

    if not meta.isa_tab and (meta.process_parameters or meta.measurement_results):
        meta.isa_tab = IsaTabRecord(
            investigation_id=path.stem,
            study_id=f"{path.stem}_text_study",
            assay_id=f"{path.stem}_text_assay",
            factor_names=list(meta.process_parameters.keys())[:6],
            factor_values=dict(meta.process_parameters),
            measurement_type="recovery_pct" if "recovery_pct" in meta.measurement_results else "assay",
            measurement_value=str(
                meta.measurement_results.get("recovery_pct")
                or meta.measurement_results.get("Cu_recovery")
                or ""
            ),
            unit="%",
        )

    return meta


def _guess_material_type(text: str) -> str | None:
    lowered = text.lower()
    for token in ("хвост", "руда", "концентрат", "шлам", "пульпа"):
        if token in lowered:
            return token
    return None


def _extract_sample_id(text: str) -> str | None:
    match = re.search(r"(?:проба|образец|sample)\s*[#№]?\s*(\w+)", text, re.I)
    return match.group(1) if match else None
