"""Универсальный парсинг таблиц Excel — любая раскладка колонок и листов."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

# KPI: ищем подстроку в любой ячейке строки (не привязка к колонке A/B)
_KPI_RULES: tuple[tuple[re.Pattern[str], str], ...] = (
    (re.compile(r"отвальн\w*\s+хвост", re.I), "tailings"),
    (re.compile(r"итого\s+извлекаем\w*\s+металл", re.I), "recoverable"),
    (re.compile(r"класс\s+крупност", re.I), "grain_header"),
)

_GRAIN_CLASS_RE = re.compile(
    r"^[+\-]?\d+(?:\s*\+\s*\d+)?(?:\s*мкм)?$",
    re.I,
)


def cell_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if abs(value - round(value)) < 1e-6:
            return str(int(round(value)))
        return f"{value:.4g}"
    text = str(value).strip()
    if text.startswith("#") and "REF" in text.upper():
        return ""
    return text


def trim_row(row: list[str]) -> list[str]:
    """Убирает пустые ячейки справа — компактный вывод для любых таблиц."""
    last = -1
    for i, cell in enumerate(row):
        if cell.strip():
            last = i
    return row[: last + 1] if last >= 0 else []


def _is_numeric_cell(text: str) -> bool:
    t = text.strip()
    if not t:
        return False
    return bool(re.fullmatch(r"[\d.,eE+\-]+", t))


def _is_text_label(text: str) -> bool:
    t = text.strip()
    if not t or len(t) < 2:
        return False
    if _is_numeric_cell(t):
        return False
    if _GRAIN_CLASS_RE.match(t):
        return False
    return bool(re.search(r"[а-яёa-zA-Z]", t))


def text_cells_in_row(row: list[str], *, limit: int = 8) -> list[tuple[int, str]]:
    """Все текстовые ячейки строки с индексами."""
    found: list[tuple[int, str]] = []
    for i, cell in enumerate(row[:limit]):
        c = cell.strip()
        if _is_text_label(c):
            found.append((i, c))
    return found


def numeric_cells_after(row: list[str], start_idx: int, *, limit: int = 6) -> list[str]:
    values: list[str] = []
    for cell in row[start_idx + 1 :]:
        c = cell.strip()
        if c and (_is_numeric_cell(c) or re.search(r"\d", c)):
            values.append(c)
            if len(values) >= limit:
                break
    return values


def grain_class_in_row(row: list[str]) -> tuple[str, int]:
    """Класс крупности в любой колонке: +71, -10, -125 +71 мкм."""
    for i, cell in enumerate(row):
        c = cell.strip()
        if not c:
            continue
        if _GRAIN_CLASS_RE.match(c) or re.match(r"^[+\-]\d", c):
            return c, i
    return "", -1


def match_row_rule(row: list[str]) -> tuple[str, str, int] | None:
    """
    Возвращает (rule_id, matched_label, cell_index) если строка подходит под KPI-правило.
    Сканирует все ячейки — работает при пустой колонке A, B или сдвинутой таблице.
    """
    for i, cell in enumerate(row):
        cl = cell.strip().lower()
        if not cl:
            continue
        for pattern, rule_id in _KPI_RULES:
            if pattern.search(cl):
                return rule_id, cell.strip(), i
    return None


def detect_header_row(rows: list[list[str]], *, scan_limit: int = 25) -> tuple[int, list[str]]:
    """Первая строка, похожая на заголовок таблицы (для произвольных xlsx)."""
    best_idx = 0
    best_score = 0
    for i, row in enumerate(rows[:scan_limit]):
        trimmed = trim_row(row)
        if len(trimmed) < 2:
            continue
        text_count = sum(1 for c in trimmed if _is_text_label(c))
        num_count = sum(1 for c in trimmed if _is_numeric_cell(c))
        score = text_count * 2 - num_count
        if text_count >= 2 and score > best_score:
            best_score = score
            best_idx = i
    header = trim_row(rows[best_idx]) if rows else []
    return best_idx, header


def summarize_sheet_domain(rows: list[list[str]]) -> list[str]:
    """Доменная сводка (хвосты, крупность) — если паттерны найдены в листе."""
    lines: list[str] = []
    tailings_tonnage: str | None = None
    recoverable_pct: str | None = None
    grain_rows: list[str] = []
    in_grain = False

    for row in rows:
        trimmed = trim_row(row)
        if not trimmed:
            continue

        matched = match_row_rule(trimmed)
        if matched:
            rule_id, label, idx = matched
            if rule_id == "tailings":
                vals = numeric_cells_after(trimmed, idx)
                if vals:
                    tailings_tonnage = vals[0]
            elif rule_id == "recoverable":
                vals = numeric_cells_after(trimmed, idx)
                for v in vals[:4]:
                    if re.search(r"\d", v):
                        recoverable_pct = v
                        break
            elif rule_id == "grain_header":
                in_grain = True
            continue

        if in_grain:
            grain_label, gidx = grain_class_in_row(trimmed)
            if grain_label:
                vals = numeric_cells_after(trimmed, gidx)
                share = vals[0] if vals else ""
                recoverable = vals[1] if len(vals) > 1 else ""
                display = grain_label if "мкм" in grain_label.lower() else f"{grain_label} мкм"
                grain_rows.append(
                    f"- {display}: доля класса {share}%, извлекаемый металл {recoverable}%"
                )
                continue
            # выход из секции крупности
            labels = [t[1].lower() for t in text_cells_in_row(trimmed, limit=3)]
            if any(lbl.startswith("итого") or "раскрыт" in lbl for lbl in labels):
                in_grain = False

    if tailings_tonnage:
        lines.append(f"- Отвальные хвосты: {tailings_tonnage} т")
    if recoverable_pct:
        lines.append(f"- Итого извлекаемый металл в хвостах: {recoverable_pct}%")
    lines.extend(grain_rows[:10])
    return lines


def summarize_sheet_generic(sheet_name: str, rows: list[list[str]]) -> list[str]:
    """Fallback-сводка для любого листа без доменных KPI."""
    if not rows:
        return [f"- Лист «{sheet_name}»: пустой"]

    non_empty = [trim_row(r) for r in rows if any(c.strip() for c in r)]
    if not non_empty:
        return [f"- Лист «{sheet_name}»: пустой"]

    max_cols = max(len(r) for r in non_empty)
    header_idx, headers = detect_header_row(non_empty)
    lines = [
        f"- Лист «{sheet_name}»: {len(non_empty)} строк × {max_cols} колонок",
    ]
    if headers:
        cols = [h for h in headers if h.strip()][:8]
        if cols:
            lines.append(f"  Колонки: {', '.join(cols)}")

    # Примеры строк «метка: значения» из данных после заголовка
    samples = 0
    for row in non_empty[header_idx + 1 :]:
        if samples >= 5:
            break
        texts = text_cells_in_row(row, limit=4)
        if not texts:
            continue
        _, label = texts[0]
        idx = texts[0][0]
        vals = numeric_cells_after(row, idx, limit=4)
        if vals:
            lines.append(f"  · {label}: {', '.join(vals)}")
            samples += 1

    return lines


def build_workbook_summary(
    sheets: dict[str, list[list[str]]],
    *,
    filename: str,
) -> list[str]:
    """Сводка по всей книге: доменные KPI + generic fallback по листам."""
    all_rows: list[list[str]] = []
    for rows in sheets.values():
        all_rows.extend(rows)

    lines: list[str] = [f"- Файл: {filename}, листов: {len(sheets)}"]

    domain = summarize_sheet_domain(all_rows)
    if domain:
        lines.append("## Сводка KPI")
        lines.extend(domain)
    else:
        lines.append("## Обзор листов")
        for name, rows in sheets.items():
            lines.extend(summarize_sheet_generic(name, rows))

    return lines


def format_sheet_rows(sheet_name: str, rows: list[list[str]]) -> list[str]:
    parts = [f"## Sheet: {sheet_name}"]
    for row in rows:
        trimmed = trim_row(row)
        if trimmed:
            parts.append(" | ".join(trimmed))
    return parts


def read_xlsx_sheets(path: Path) -> dict[str, list[list[str]]]:
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    sheets: dict[str, list[list[str]]] = {}
    for sheet in wb.worksheets:
        rows: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            cells = [cell_str(c) for c in row]
            if any(cells):
                rows.append(cells)
        sheets[sheet.title] = rows
    return sheets


def parse_spreadsheet(path: Path) -> str:
    """Точка входа: любой xlsx/xls → текст для RAG."""
    sheets = read_xlsx_sheets(path)
    summary = build_workbook_summary(sheets, filename=path.name)

    parts = [f"# Данные Excel: {path.name}", "## Сводка для RAG", *summary, ""]
    for name, rows in sheets.items():
        parts.extend(format_sheet_rows(name, rows))
    return "\n".join(parts)


# Обратная совместимость для тестов
def parse_xlsx_summary(rows: list[list[str]]) -> list[str]:
    domain = summarize_sheet_domain(rows)
    if domain:
        return domain
    return summarize_sheet_generic("sheet", rows)
