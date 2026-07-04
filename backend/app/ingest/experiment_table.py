"""Таблица лабораторных экспериментов Excel: одна строка = один опыт."""

from __future__ import annotations

import hashlib
import re
from pathlib import Path
from typing import Any

from app.config import settings
from app.models import DocumentMetadata, IsaTabRecord

# Канонические поля для ML (см. predictive_model._FEATURE_KEYS)
_CANONICAL_FIELDS = ("sample_id", "pH", "reagent_dosage", "temperature_C", "recovery_pct")

# Подстроки в заголовке колонки → каноническое имя
_HEADER_RULES: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("sample_id", ("sample_id", "sample", "проба", "образец", "id пробы")),
    ("pH", ("ph",)),
    ("reagent_dosage", ("reagent_dosage", "dosage", "доз", "реагент", "kg_t", "кг/т", "кг_т")),
    ("temperature_C", ("temperature", "temp", "темп")),
    ("recovery_pct", ("recovery", "recover", "извлеч", "выход", "содерж")),
)


def _normalize_header(value: object) -> str:
    text = str(value or "").strip().lower()
    text = text.replace(" ", "_").replace("-", "_")
    return text


def _canonical_column(header: str) -> str | None:
    h = _normalize_header(header)
    if not h:
        return None
    for canonical, patterns in _HEADER_RULES:
        if any(p in h for p in patterns):
            return canonical
    return None


def map_experiment_columns(headers: list[object]) -> dict[str, int]:
    """Каноническое имя поля → индекс колонки."""
    mapped: dict[str, int] = {}
    for i, header in enumerate(headers):
        canonical = _canonical_column(str(header) if header is not None else "")
        if canonical and canonical not in mapped:
            mapped[canonical] = i
    return mapped


def is_experiment_table(headers: list[object]) -> bool:
    """Таблица опытов: ≥2 фактора (pH/доза/темп) + целевой показатель извлечения."""
    mapped = map_experiment_columns(headers)
    factor_count = sum(1 for key in ("pH", "reagent_dosage", "temperature_C") if key in mapped)
    return factor_count >= 2 and "recovery_pct" in mapped


def _parse_number(value: object) -> float | str | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    match = re.search(r"(\d+(?:\.\d+)?)", text)
    return float(match.group(1)) if match else text


def _row_has_data(row: tuple[object, ...] | list[object], col_map: dict[str, int]) -> bool:
    recovery_idx = col_map.get("recovery_pct")
    if recovery_idx is None or recovery_idx >= len(row):
        return False
    val = _parse_number(row[recovery_idx])
    return isinstance(val, (int, float))


def parse_experiment_rows(
    rows: list[tuple[object, ...] | list[object]],
    *,
    header_idx: int = 0,
) -> list[dict[str, Any]] | None:
    if len(rows) <= header_idx + 1:
        return None

    headers = rows[header_idx]
    if not is_experiment_table(headers):
        return None

    col_map = map_experiment_columns(headers)
    raw_headers = [
        str(headers[i]).strip() if i < len(headers) and headers[i] is not None else ""
        for i in range(len(headers))
    ]

    records: list[dict[str, Any]] = []
    for row in rows[header_idx + 1 :]:
        if not _row_has_data(row, col_map):
            continue

        record: dict[str, Any] = {"raw": {}}
        for canonical, idx in col_map.items():
            if idx >= len(row):
                continue
            cell = row[idx]
            if cell is None or str(cell).strip() == "":
                continue
            if canonical == "sample_id":
                record["sample_id"] = str(cell).strip()
            else:
                parsed = _parse_number(cell)
                if parsed is not None:
                    record[canonical] = parsed

            header_name = raw_headers[idx] if idx < len(raw_headers) else str(idx)
            if header_name:
                record["raw"][header_name] = cell

        if "recovery_pct" not in record:
            continue
        if "sample_id" not in record:
            record["sample_id"] = f"row_{len(records) + 1}"
        records.append(record)

    return records or None


def parse_experiment_table(path: Path) -> list[dict[str, Any]] | None:
    """Читает xlsx и возвращает список экспериментов или None для обычных таблиц."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None

    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
    except Exception:
        return None

    sheet = wb.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))[:500]
    if len(rows) < 2:
        return None

    # Сначала первая строка (типичный формат), иначе эвристика по текстовым заголовкам
    result = parse_experiment_rows(rows, header_idx=0)
    if result:
        return result

    for i, row in enumerate(rows[:15]):
        if is_experiment_table(row):
            return parse_experiment_rows(rows, header_idx=i)

    return None


def experiment_row_to_text(path: Path, row: dict[str, Any]) -> str:
    sample = row.get("sample_id", "?")
    lines = [f"# Лабораторный опыт {sample} ({path.name})"]
    if "pH" in row:
        lines.append(f"- pH: {row['pH']}")
    if "reagent_dosage" in row:
        lines.append(f"- Дозировка реагента: {row['reagent_dosage']} кг/т")
    if "temperature_C" in row:
        lines.append(f"- Температура: {row['temperature_C']} °C")
    if "recovery_pct" in row:
        lines.append(f"- Извлечение: {row['recovery_pct']}%")
    return "\n".join(lines)


def _relative_source(path: Path) -> str:
    try:
        return str(path.relative_to(settings.data_dir_path))
    except ValueError:
        return str(path)


def build_experiment_row_metadata(
    path: Path,
    row: dict[str, Any],
    base: DocumentMetadata | None = None,
) -> DocumentMetadata:
    sample = str(row.get("sample_id", "row"))
    rel = _relative_source(path)
    meta = (base.model_copy(deep=True) if base else DocumentMetadata())
    meta.source = f"{rel}#{sample}"
    meta.title = f"{path.stem} / {sample}"
    meta.sample_id = sample
    meta.tags = list(dict.fromkeys(meta.tags + ["structured_experiment", "experiment_row", "xlsx_type:lab_experiments"]))
    meta.material_type = meta.material_type or "лабораторный опыт"

    for key in ("pH", "reagent_dosage", "temperature_C"):
        if key in row:
            meta.process_parameters[key] = row[key]
            meta.experiment_conditions[key] = row[key]

    if "recovery_pct" in row:
        meta.measurement_results["recovery_pct"] = row["recovery_pct"]

    meta.isa_tab = IsaTabRecord(
        investigation_id=path.stem,
        study_id=f"{path.stem}_study",
        assay_id=f"{path.stem}_{sample}",
        factor_names=[k for k in ("pH", "reagent_dosage", "temperature_C") if k in row],
        factor_values={k: row[k] for k in ("pH", "reagent_dosage", "temperature_C") if k in row},
        measurement_type="recovery_pct",
        measurement_value=str(row.get("recovery_pct", "")),
        unit="%",
        raw_sample_characteristics=row.get("raw") or {},
    )
    meta.allotrope_process_uri = meta.allotrope_process_uri or (
        f"allotrope:process/flotation/{path.stem}-laboratory-assay"
    )
    return meta


def make_experiment_row_doc_id(path: Path, row: dict[str, Any]) -> str:
    sample = str(row.get("sample_id", "row"))
    key = "|".join(
        str(row.get(field, ""))
        for field in _CANONICAL_FIELDS
    )
    digest = hashlib.sha256(key.encode("utf-8")).hexdigest()[:8]
    safe = re.sub(r"[^\w\-]+", "_", sample)[:24]
    return f"{path.stem}_{safe}_{digest}"
