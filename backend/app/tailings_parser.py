"""Парсер отчётов по хвостам + детерминированный диагностический слой.

Формат входа — Excel «Хвосты *.xlsx» Норникеля (см. «Как читать отчет института
по хвостам.docx»): гранулометрия потерь по классам крупности и минералогия
внутри каждого класса. Элемент 28 = Ni, Элемент 29 = Cu.

Диагностика — правила, не LLM: по каждому классу определяем, где потери и в
какой форме, и относим к типу проблемы. LLM получает готовый диагноз.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field, asdict

import openpyxl

COARSE_CLASSES = ["+125", "-125+71", "+71", "-71+45"]
MIDDLE_CLASSES = ["-45+20"]
FINE_CLASSES = ["-20+10", "-10"]
ALL_CLASSES = COARSE_CLASSES + MIDDLE_CLASSES + FINE_CLASSES

ELEMENT_NAMES = {"28": "Ni (элемент 28)", "29": "Cu (элемент 29)"}

# Формы потерь: извлекаемые текущей технологией и нет (по инструкции института)
EXTRACTABLE_FORMS = {"раскрытый", "закрытый", "миллерит"}
FORM_PATTERNS = [
    ("раскрытый", re.compile(r"раскрыт", re.I)),
    ("закрытый", re.compile(r"закрыт", re.I)),
    ("примесь в пирротине", re.compile(r"пирротин", re.I)),
    ("силикатная форма", re.compile(r"силикат|валлериит", re.I)),
    ("пирит/прочие сульфиды", re.compile(r"пирит|другие", re.I)),
    ("миллерит", re.compile(r"миллерит", re.I)),
]

PROBLEM_TYPES = {
    "regrind": "Недораскрытие: сростки в крупных классах — нужны доизмельчение и/или улучшение классификации",
    "coarse_flotation": "Раскрытые зёрна теряются в крупных классах — флотация не удерживает крупные частицы (время флотации, фронт, реагентный режим)",
    "slimes": "Шламовые потери: тонкие классы (-20 мкм) не флотируются — переизмельчение и/или потери тонких раскрытых частиц",
    "not_extractable": "Технологически неизвлекаемо текущей схемой (силикатные формы, примесь в решётке пирротина)",
}


def _norm_class(label: str) -> str | None:
    """' -20 + 10 мкм' → '-20+10'"""
    s = re.sub(r"мкм", "", str(label)).replace(" ", "").strip()
    return s if s in ALL_CLASSES else None


def _num(v) -> float | None:
    if v is None or (isinstance(v, str) and ("#REF" in v or not v.strip())):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _form_of(label: str) -> str | None:
    for form, pat in FORM_PATTERNS:
        if pat.search(str(label)):
            return form
    return None


@dataclass
class ClassLoss:
    size_class: str
    share_pct: float | None = None          # доля класса в массе хвостов
    loss: dict = field(default_factory=dict)     # {"28": {"share_pct":..,"tons":..}, "29": ...}
    mineralogy: dict = field(default_factory=dict)  # {"28": {форма: {"share_pct":..,"tons":..}}, ...}
    extractable: dict = field(default_factory=dict)  # {"28": {"share_pct":..,"tons":..}, ...}


@dataclass
class Section:
    name: str                                # "Хвосты породные" / "Хвосты пирротиновые"
    mass_smt: float | None = None
    totals: dict = field(default_factory=dict)   # {"28": {"pct":..,"tons":..}, "29": ...}
    classes: list[ClassLoss] = field(default_factory=list)


def parse_tailings_xlsx(path: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [[c for c in row] for row in ws.iter_rows(values_only=True)]

    sections: list[Section] = []
    current: Section | None = None
    current_class: ClassLoss | None = None
    in_class_table = False

    for row in rows:
        label = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        if not label and not any(v is not None for v in row):
            continue
        cells = list(row) + [None] * (8 - len(row))

        # начало секции: "Хвосты породные" / "Хвосты пирротиновые" с массой
        if (
            re.match(r"^Хвосты\s+(породные|пирротиновые)", label, re.I)
            and _num(cells[2])
            and (_num(cells[4]) is not None or _num(cells[6]) is not None)
        ):
            current = Section(
                name=label,
                mass_smt=_num(cells[2]),
                totals={
                    "28": {"pct": _num(cells[3]), "tons": _num(cells[4])},
                    "29": {"pct": _num(cells[5]), "tons": _num(cells[6])},
                },
            )
            sections.append(current)
            in_class_table = False
            current_class = None
            continue

        if current is None:
            continue

        # таблица гранулометрии
        if label.startswith("Класс крупности"):
            in_class_table = True
            continue
        if in_class_table:
            cls = _norm_class(label)
            if cls:
                current.classes.append(
                    ClassLoss(
                        size_class=cls,
                        share_pct=_num(cells[2]),
                        loss={
                            "28": {"share_pct": _num(cells[3]), "tons": _num(cells[4])},
                            "29": {"share_pct": _num(cells[5]), "tons": _num(cells[6])},
                        },
                    )
                )
                continue
            if label.startswith("Итого"):
                in_class_table = False
                continue

        # минералогический блок: заголовок = имя класса с "мкм"
        cls = _norm_class(label)
        if cls and "мкм" in label:
            current_class = next((c for c in current.classes if c.size_class == cls), None)
            if current_class is not None:
                current_class.mineralogy = {"28": {}, "29": {}}
            continue

        if current_class is not None:
            form = _form_of(label)
            if form:
                for el, (pcol, tcol) in {"28": (3, 4), "29": (5, 6)}.items():
                    pct, tons = _num(cells[pcol]), _num(cells[tcol])
                    if pct is not None or tons is not None:
                        current_class.mineralogy[el][form] = {"share_pct": pct, "tons": tons}
                continue
            if label.startswith("Извлекаемый"):
                current_class.extractable = {
                    "28": {"share_pct": _num(cells[3]), "tons": _num(cells[4])},
                    "29": {"share_pct": _num(cells[5]), "tons": _num(cells[6])},
                }
                continue
            if label.startswith("Итого") or label.startswith("Не извлекаемый"):
                continue

    # в файле могут идти подряд несколько грануло-таблиц (факт/расчёт) — берём первую
    for s in sections:
        seen: set[str] = set()
        s.classes = [c for c in s.classes if not (c.size_class in seen or seen.add(c.size_class))]

    result = {
        "file": path.split("/")[-1],
        "sections": [
            {**asdict(s), "classes": [asdict(c) for c in s.classes]} for s in sections
        ],
    }
    result["diagnostics"] = diagnose(result)
    result["summary_text"] = build_summary_text(result)
    return result


def diagnose(parsed: dict) -> list[dict]:
    """Правила: класс × элемент → тип проблемы, приоритет по извлекаемым тоннам."""
    diags: list[dict] = []
    for sec in parsed["sections"]:
        for cls in sec["classes"]:
            size = cls["size_class"]
            for el in ("28", "29"):
                loss_tons = (cls["loss"].get(el) or {}).get("tons")
                if not loss_tons or loss_tons <= 0:
                    continue
                minerals = (cls.get("mineralogy") or {}).get(el) or {}
                extractable_tons = (cls.get("extractable") or {}).get(el, {}).get("tons")

                opened = (minerals.get("раскрытый") or {}).get("tons") or 0
                locked = (minerals.get("закрытый") or {}).get("tons") or 0
                millerite = (minerals.get("миллерит") or {}).get("tons") or 0

                if minerals:
                    known = {k: (v.get("tons") or 0) for k, v in minerals.items()}
                    dominant_form = max(known, key=known.get) if known else None
                else:
                    dominant_form = None

                # приоритет извлекаемого: если минералогии нет — считаем по валовым потерям
                target_tons = extractable_tons if extractable_tons is not None else loss_tons

                if size in COARSE_CLASSES or size in MIDDLE_CLASSES:
                    if minerals and locked >= opened + millerite:
                        problem = "regrind"
                    elif minerals:
                        problem = "coarse_flotation"
                    else:
                        problem = "regrind"  # без минералогии в крупном классе — базовая гипотеза
                else:
                    problem = "slimes"

                if minerals and dominant_form not in EXTRACTABLE_FORMS and target_tons < loss_tons * 0.3:
                    problem = "not_extractable"

                diags.append(
                    {
                        "section": sec["name"],
                        "size_class": size,
                        "element": el,
                        "element_name": ELEMENT_NAMES[el],
                        "loss_tons": round(loss_tons, 1),
                        "extractable_tons": round(target_tons, 1) if target_tons else None,
                        "dominant_form": dominant_form,
                        "opened_tons": round(opened, 1),
                        "locked_tons": round(locked, 1),
                        "problem_type": problem,
                        "problem_text": PROBLEM_TYPES[problem],
                    }
                )
    # приоритет: сколько извлекаемого металла «лежит» в этой ячейке
    diags.sort(key=lambda d: -(d["extractable_tons"] or 0))
    return diags


def build_summary_text(parsed: dict) -> str:
    """Компактный текст диагноза для промпта LLM (и для показа пользователю)."""
    lines: list[str] = []
    for sec in parsed["sections"]:
        t28 = (sec["totals"].get("28") or {}).get("tons")
        t29 = (sec["totals"].get("29") or {}).get("tons")
        lines.append(
            f"### {sec['name']}: {sec['mass_smt']:,.0f} т, потери Ni {t28 or 0:,.0f} т, Cu {t29 or 0:,.0f} т".replace(",", " ")
        )
        for cls in sec["classes"]:
            l28 = (cls["loss"].get("28") or {}).get("share_pct")
            l29 = (cls["loss"].get("29") or {}).get("share_pct")
            parts = [f"класс {cls['size_class']} мкм: {cls['share_pct'] or 0:.1f}% массы"]
            if l28 is not None:
                parts.append(f"{l28:.1f}% потерь Ni")
            if l29 is not None:
                parts.append(f"{l29:.1f}% потерь Cu")
            lines.append("  - " + ", ".join(parts))

    top = [d for d in parsed["diagnostics"] if d["problem_type"] != "not_extractable"][:8]
    if top:
        lines.append("\n### Диагноз (главные адреса извлекаемых потерь):")
        for d in top:
            form = f", доминирует форма: {d['dominant_form']}" if d["dominant_form"] else ""
            lines.append(
                f"  - {d['element_name']}, класс {d['size_class']} мкм ({d['section']}): "
                f"потери {d['loss_tons']} т, извлекаемо ~{d['extractable_tons'] or '?'} т{form}. "
                f"{d['problem_text']}"
            )
    return "\n".join(lines)
